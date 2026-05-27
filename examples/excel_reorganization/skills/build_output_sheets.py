"""Create Excel workbook with monthly sheets and subtotals.

This skill creates an Excel workbook with one sheet per month (YYYY-MM format),
populates each sheet with sales data rows sorted by employee name, and adds
a subtotal row at the bottom of each sheet.

Pattern: Follows examples/rpa_challenge/skills/setup.py:95-112
"""

from __future__ import annotations
from pathlib import Path
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from oref import BusinessException, ProcessContext, Skill, SystemException, get_logger

logger = get_logger(__name__)

class BuildOutputSheets(Skill):
    """Create Excel workbook with monthly sheets and subtotals."""

    def execute(self, ctx: ProcessContext) -> None:
        """Create workbook with monthly sheets, populate data, and save to file."""
        grouped_data = ctx.data.get("grouped_data")
        output_dir = ctx.data.get("output_dir")
        output_filename = ctx.data.get("output_filename")

        if grouped_data is None:
            raise BusinessException(
                "No grouped_data in context — GroupByMonth must run before this skill",
                action=self.name,
            )
        if output_dir is None:
            raise BusinessException("No output_dir in context.", action=self.name)
        if output_filename is None:
            raise BusinessException("No output_filename in context.", action=self.name)
        if not grouped_data:
            raise BusinessException("No grouped data available.", action=self.name)

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Create workbook with one sheet per month
        # Use write_only=True to avoid creating default sheet
        workbook = Workbook(write_only=True)

        # Create sheets and populate data
        for year_month, month_data in sorted(grouped_data.items()):
            sheet_name = year_month
            ws = workbook.create_sheet(sheet_name)

            # Write header row
            header_row = ["Employee Name", "Date", "Amount", "Country"]
            ws.append(header_row)

            # Write data rows
            for row in month_data:
                # Validate employee name is not empty
                emp_name = row.get("employee_name")
                if not emp_name:
                    raise BusinessException(
                        f"Row has empty employee name",
                        action=self.name,
                    )
                ws.append([
                    emp_name,
                    row.get("date"),
                    row.get("amount"),
                    row.get("country"),
                ])

            # Calculate and write subtotal (verify amounts are numeric)
            if not month_data:
                continue  # Skip if no data
            subtotal = sum(float(row["amount"]) for row in month_data)
            ws.append(["Subtotal", "", subtotal, ""])

        output_filename = self._resolve_output_filename(str(output_filename), grouped_data)
        output_path = output_dir / output_filename

        try:
            # Save workbook with write_only mode.
            workbook.save(output_path)

            # Load the workbook again to apply formatting. This is necessary because
            # WriteOnlyWorksheet doesn't support cell() method.
            workbook = load_workbook(output_path)

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                self._apply_sheet_formatting(ws)

            workbook.save(output_path)
        except Exception as exc:
            if output_path.exists():
                output_path.unlink()
            raise SystemException(f"Failed to build Excel workbook: {exc}", action=self.name) from exc

        logger.info("Created Excel workbook: %s", output_path)

        # Store output path in context for VerifyOutput skill
        ctx.data["output_path"] = str(output_path)

    def _apply_sheet_formatting(self, ws: Any) -> None:
        """Apply basic formatting to worksheet."""
        # Header row formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(["Employee Name", "Date", "Amount", "Country"], 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Subtotal row formatting
        subtotal_font = Font(bold=True)
        # Count actual data rows (excluding header and subtotal)
        data_rows = [row for row in ws.iter_rows(values_only=True) 
                     if row[0] not in ["Employee Name", "Subtotal"]]
        
        # Subtotal is at header + data rows + 1.
        subtotal_row = len(data_rows) + 2
        
        for col_idx in range(1, 5):
            cell = ws.cell(row=subtotal_row, column=col_idx)
            cell.font = subtotal_font

    def _resolve_output_filename(self, output_filename: str, grouped_data: dict[str, list[dict[str, Any]]]) -> str:
        """Resolve configured output filename, allowing month placeholders."""
        if "{month}" not in output_filename:
            return output_filename

        first_date = min(
            row["date"]
            for month_data in grouped_data.values()
            for row in month_data
        )
        return output_filename.format(month=first_date[:7])
