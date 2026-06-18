"""Verify output Excel file is readable and complete.

This skill verifies that the output Excel file can be read with openpyxl,
contains all expected months, and has the correct structure (header + data + subtotal).
It also verifies Transaction.metadata contains required fields.

Pattern: Follows examples/json_event_log_processor/skills/validate_events.py:3-64
"""

from __future__ import annotations
import math
from pathlib import Path
from openpyxl import load_workbook
from rpacore import BusinessException, ProcessContext, Skill, get_logger

logger = get_logger(__name__)

class VerifyOutput(Skill):
    """Verify output Excel file is readable and complete."""

    def execute(self, ctx: ProcessContext) -> None:
        """Verify output file exists, is readable, has correct structure, and metadata is set."""
        output_path = ctx.require_state("output_path", str, action=self.name)
        expected_months = ctx.require_state("expected_months", list, action=self.name)
        grouped_data = ctx.require_state("grouped_data", dict, action=self.name)
        expected_subtotals = ctx.require_state("expected_subtotals", dict, action=self.name)

        output_path = Path(output_path)

        # Check file exists
        if not output_path.exists():
            self._fail(f"Output file does not exist: {output_path}")

        # Verify file is readable
        try:
            workbook = load_workbook(output_path)
        except Exception as exc:
            raise BusinessException(f"Cannot read output file: {exc}", action=self.name, stop=True) from exc

        # Verify sheet names match expected months
        actual_months = list(workbook.sheetnames)
        expected_month_set = set(expected_months)
        missing_months = expected_month_set - set(actual_months)
        extra_months = set(actual_months) - expected_month_set

        if missing_months:
            self._fail(f"Missing months in output: {missing_months}")
        if extra_months:
            self._fail(f"Unexpected months in output: {extra_months}")

        for month in expected_months:
            ws = workbook[month]
            rows = list(ws.iter_rows(values_only=True))
            month_data = grouped_data.get(month, [])

            expected_rows = len(month_data) + 2
            if len(rows) != expected_rows:
                self._fail(f"Sheet '{month}' has {len(rows)} rows; expected {expected_rows}")

            if tuple(rows[0]) != ("Employee Name", "Date", "Amount", "Country"):
                self._fail(f"Sheet '{month}' has incorrect header: {rows[0]}")

            # Verify subtotal row exists
            if rows[-1][0] != "Subtotal":
                self._fail(f"Sheet '{month}' missing subtotal row")

            data_rows = rows[1:-1]
            expected_names = [row["employee_name"] for row in month_data]
            actual_names = [row[0] for row in data_rows]
            if actual_names != expected_names:
                self._fail(f"Sheet '{month}' has incorrect employee ordering")

            expected_subtotal = expected_subtotals.get(month)
            if expected_subtotal is None:
                self._fail(f"Missing expected subtotal for sheet '{month}'")
            actual_subtotal = rows[-1][2]
            if not math.isclose(float(actual_subtotal), float(expected_subtotal)):
                self._fail(
                    f"Sheet '{month}' subtotal {actual_subtotal} does not match expected {expected_subtotal}"
                )

        # Verify Transaction.metadata contains required fields
        metadata = ctx.transaction.metadata
        required_meta_keys = {
            "source_csv", "row_count", "month_count",
            "output_path", "employee_count", "generated_at",
        }
        missing_meta = required_meta_keys - set(metadata.keys())
        if missing_meta:
            self._fail(f"Transaction.metadata missing required fields: {missing_meta}")

        # Verify metadata values are consistent with actual output
        if metadata.get("month_count") != len(expected_months):
            self._fail(
                f"Metadata month_count {metadata.get('month_count')} does not match actual {len(expected_months)}"
            )
        if metadata.get("output_path") != str(output_path):
            self._fail(f"Metadata output_path {metadata.get('output_path')} does not match actual {output_path}")

        logger.info("Output verification passed: %s", output_path)

    def _fail(self, message: str) -> None:
        """Raise a permanent verification failure."""
        raise BusinessException(message, action=self.name, stop=True)
