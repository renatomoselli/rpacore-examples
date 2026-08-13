"""Create Excel workbook with monthly sheets and subtotals.

This step creates an Excel workbook with one sheet per month (YYYY-MM format),
populates each sheet with sales data rows sorted by employee name, and adds
a subtotal row at the bottom of each sheet.

Pattern: Follows examples/json_event_log_processor/steps/write_output.py:39-45
"""

from __future__ import annotations
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from rpacore import (
    BusinessException,
    ProcessContext,
    Step,
    SystemException,
    atomic_output_path,
    get_logger,
)

logger = get_logger(__name__)

class BuildOutputSheets(Step):
    """Create Excel workbook with monthly sheets and subtotals."""

    def execute(self, ctx: ProcessContext) -> None:
        """Create workbook with monthly sheets, populate data, and save to file."""
        grouped_data = ctx.require_state("grouped_data", dict, action=self.name)
        output_dir = ctx.require_config("output_dir", str, action=self.name)
        output_filename = ctx.require_state("output_filename", str, action=self.name)

        if not grouped_data:
            raise BusinessException("No grouped data available.", action=self.name)

        for month_data in grouped_data.values():
            for row in month_data:
                if not row.get("employee_name"):
                    raise BusinessException(
                        "Row has empty employee name",
                        action=self.name,
                    )

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        output_filename = self._resolve_output_filename(str(output_filename), grouped_data)
        output_path = output_dir / output_filename
        output_dir_resolved = output_dir.resolve()
        output_path_resolved = output_path.resolve()
        if not output_path_resolved.is_relative_to(output_dir_resolved):
            raise SystemException(
                f"Output path escapes output_dir: {output_path}",
                action=self.name,
            )
        output_path = output_path_resolved

        # Create workbook with one sheet per month
        # Use write_only=True to avoid creating default sheet
        workbook = Workbook(write_only=True)

        # Create sheets and populate data
        for year_month, month_data in sorted(grouped_data.items()):
            sheet_name = year_month
            ws = workbook.create_sheet(sheet_name)

            # Write header row
            header_row = ["Employee Name", "Date", "Amount", "Country"]
            ws.append(header_row)

            # Write data rows
            for row in month_data:
                emp_name = row.get("employee_name")
                ws.append([
                    emp_name,
                    row.get("date"),
                    row.get("amount"),
                    row.get("country"),
                ])

            # Calculate and write subtotal (verify amounts are numeric)
            subtotal = sum(float(row["amount"]) for row in month_data)
            ws.append(["Subtotal", "", subtotal, ""])

        try:
            with atomic_output_path(output_path) as temporary_path:
                # Save workbook with write_only mode.
                workbook.save(temporary_path)

                # WriteOnlyWorksheet cannot format cells, so reopen the same
                # unpublished sibling before formatting and final publication.
                # ``atomic_output_path`` uses a .tmp suffix; openpyxl accepts
                # the sibling through a binary handle without inspecting it.
                with temporary_path.open("rb") as workbook_file:
                    workbook = load_workbook(workbook_file)

                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    self._apply_sheet_formatting(ws, data_row_count=len(grouped_data[sheet_name]))

                workbook.save(temporary_path)
        except Exception as exc:
            raise SystemException(f"Failed to build Excel workbook: {exc}", action=self.name) from exc

        logger.info("Created Excel workbook: %s", output_path)

        # Store output path in context for VerifyOutput step
        ctx.state["output_path"] = str(output_path)

        # Record artifact with metadata
        total_rows = sum(len(month_data) for month_data in grouped_data.values())
        unique_employees = len({
            row["employee_name"]
            for month_data in grouped_data.values()
            for row in month_data
        })
        generated_at = datetime.now(timezone.utc).isoformat()
        artifact_metadata = {
            "row_count": total_rows,
            "month_count": len(grouped_data),
            "employee_count": unique_employees,
            "generated_at": generated_at,
        }
        ctx.add_artifact(
            name=output_filename,
            path=str(output_path),
            kind="output",
            metadata=artifact_metadata,
        )

        # Update Transaction.metadata with computed fields for reporting
        tx = ctx.transaction
        tx.metadata["row_count"] = total_rows
        tx.metadata["month_count"] = len(grouped_data)
        tx.metadata["output_path"] = str(output_path)
        tx.metadata["employee_count"] = unique_employees
        tx.metadata["generated_at"] = generated_at

    def _apply_sheet_formatting(self, ws: Any, *, data_row_count: int) -> None:
        """Apply basic formatting to worksheet."""
        # Header row formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(["Employee Name", "Date", "Amount", "Country"], 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Subtotal row formatting
        subtotal_font = Font(bold=True)

        # Subtotal is at header + data rows + 1.
        subtotal_row = data_row_count + 2

        for col_idx in range(1, 5):
            cell = ws.cell(row=subtotal_row, column=col_idx)
            cell.font = subtotal_font

    def _resolve_output_filename(self, output_filename: str, grouped_data: dict[str, list[dict[str, Any]]]) -> str:
        """Resolve configured output filename, allowing month placeholders."""
        if "{month}" not in output_filename:
            return output_filename

        first_date = min(
            row["date"]
            for month_data in grouped_data.values()
            for row in month_data
        )
        return output_filename.format(month=first_date[:7])
