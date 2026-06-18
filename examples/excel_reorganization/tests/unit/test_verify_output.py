from __future__ import annotations

import pytest
from openpyxl import Workbook
from rpacore import BusinessException

from skills.verify_output import VerifyOutput
from tests.conftest import make_context


def _grouped_data() -> dict:
    return {
        "2024-01": [
            {"employee_name": "Alice", "date": "2024-01-05", "amount": 10.0, "country": "USA"},
            {"employee_name": "Zoe", "date": "2024-01-10", "amount": 15.5, "country": "UK"},
        ]
    }


def _write_workbook(path, rows):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "2024-01"
    for row in rows:
        ws.append(row)
    workbook.save(path)


def _metadata(output_path, month_count=1):
    return {
        "source_csv": "sample.csv",
        "row_count": 2,
        "month_count": month_count,
        "output_path": str(output_path),
        "employee_count": 2,
        "generated_at": "2024-01-15T00:00:00+00:00",
    }


def _verify_context(output_path, *, grouped_data=None, expected_months=None, expected_subtotals=None):
    return make_context(
        state={
            "output_path": str(output_path),
            "expected_months": expected_months or ["2024-01"],
            "grouped_data": grouped_data or _grouped_data(),
            "expected_subtotals": {"2024-01": 25.5} if expected_subtotals is None else expected_subtotals,
        },
    )


def test_verify_output_accepts_complete_workbook(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Subtotal", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path)
    # Seed transaction metadata (normally set by BuildOutputSheets)
    ctx.transaction.metadata.update(_metadata(output_path))

    VerifyOutput(name="verify_output", execution_order=1).execute(ctx)


def test_verify_output_rejects_missing_data_rows(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Subtotal", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path)
    ctx.transaction.metadata.update(_metadata(output_path))

    with pytest.raises(BusinessException, match="expected 4") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_wrong_subtotal(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Subtotal", "", 99.0, ""],
        ],
    )
    ctx = _verify_context(output_path)
    ctx.transaction.metadata.update(_metadata(output_path))

    with pytest.raises(BusinessException, match="subtotal") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_missing_metadata(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Subtotal", "", 10.0, ""],
        ],
    )
    ctx = _verify_context(
        output_path,
        grouped_data={
            "2024-01": [{"employee_name": "Alice", "date": "2024-01-05", "amount": 10.0, "country": "USA"}]
        },
        expected_subtotals={"2024-01": 10.0},
    )
    # Do NOT seed metadata — should fail

    with pytest.raises(BusinessException, match="Transaction.metadata missing") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_uses_independent_expected_subtotals(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Subtotal", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path, expected_subtotals={"2024-01": 99.0})
    ctx.transaction.metadata.update(_metadata(output_path))

    with pytest.raises(BusinessException, match="subtotal") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_missing_output_file(tmp_path):
    output_path = tmp_path / "missing.xlsx"
    ctx = _verify_context(output_path)

    with pytest.raises(BusinessException, match="does not exist") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_unreadable_output_file(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    output_path.write_text("not a workbook", encoding="utf-8")
    ctx = _verify_context(output_path)

    with pytest.raises(BusinessException, match="Cannot read output file") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_missing_month(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(output_path, [["Employee Name", "Date", "Amount", "Country"], ["Subtotal", "", 0, ""]])
    ctx = _verify_context(output_path, expected_months=["2024-01", "2024-02"])
    ctx.transaction.metadata.update(_metadata(output_path, month_count=2))

    with pytest.raises(BusinessException, match="Missing months") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_extra_month(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    workbook = Workbook()
    workbook.active.title = "2024-01"
    workbook.create_sheet("2024-02")
    workbook.save(output_path)
    ctx = _verify_context(output_path)
    ctx.transaction.metadata.update(_metadata(output_path))

    with pytest.raises(BusinessException, match="Unexpected months") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_wrong_header(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Subtotal", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path)
    ctx.transaction.metadata.update(_metadata(output_path))

    with pytest.raises(BusinessException, match="incorrect header") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_missing_subtotal_row(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Total", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path)
    ctx.transaction.metadata.update(_metadata(output_path))

    with pytest.raises(BusinessException, match="missing subtotal") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_wrong_employee_order(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Subtotal", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path)
    ctx.transaction.metadata.update(_metadata(output_path))

    with pytest.raises(BusinessException, match="employee ordering") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_missing_expected_subtotal(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Subtotal", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path, expected_subtotals={})
    ctx.transaction.metadata.update(_metadata(output_path))

    with pytest.raises(BusinessException, match="Missing expected subtotal") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_wrong_metadata_month_count(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Subtotal", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path)
    ctx.transaction.metadata.update(_metadata(output_path, month_count=2))

    with pytest.raises(BusinessException, match="month_count") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True


def test_verify_output_rejects_wrong_metadata_output_path(tmp_path):
    output_path = tmp_path / "sales.xlsx"
    _write_workbook(
        output_path,
        [
            ["Employee Name", "Date", "Amount", "Country"],
            ["Alice", "2024-01-05", 10.0, "USA"],
            ["Zoe", "2024-01-10", 15.5, "UK"],
            ["Subtotal", "", 25.5, ""],
        ],
    )
    ctx = _verify_context(output_path)
    ctx.transaction.metadata.update(_metadata(tmp_path / "other.xlsx"))

    with pytest.raises(BusinessException, match="output_path") as exc_info:
        VerifyOutput(name="verify_output", execution_order=1).execute(ctx)
    assert exc_info.value.stops_execution is True
