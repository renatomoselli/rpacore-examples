from __future__ import annotations

import os

import pytest
from openpyxl import load_workbook
import skills.build_output_sheets as build_output_sheets
from rpacore import BusinessException, SystemException

from skills.build_output_sheets import BuildOutputSheets
from tests.conftest import make_context


def _grouped_data() -> dict:
    return {
        "2024-01": [
            {"employee_name": "Alice", "date": "2024-01-05", "amount": 10.0, "country": "USA"},
            {"employee_name": "Zoe", "date": "2024-01-10", "amount": 15.5, "country": "UK"},
        ],
        "2024-02": [
            {"employee_name": "Bob", "date": "2024-02-01", "amount": 7.0, "country": "Canada"},
        ],
    }


def test_build_output_sheets_creates_formatted_workbook(tmp_path):
    ctx = make_context(
        state={
            "grouped_data": _grouped_data(),
            "output_filename": "sales_report_{month}.xlsx",
        },
        config={
            "output_dir": str(tmp_path),
        },
    )

    BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)

    output_path = tmp_path / "sales_report_2024-01.xlsx"
    assert ctx.state["output_path"] == str(output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["2024-01", "2024-02"]
    ws = workbook["2024-01"]
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fill_type == "solid"
    assert ws["A4"].value == "Subtotal"
    assert ws["C4"].value == 25.5
    assert workbook["2024-02"].max_row == 3
    assert workbook["2024-02"]["A3"].font.bold is True


def test_build_output_sheets_honors_literal_output_filename(tmp_path):
    ctx = make_context(
        state={
            "grouped_data": _grouped_data(),
            "output_filename": "custom.xlsx",
        },
        config={
            "output_dir": str(tmp_path),
        },
    )

    BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)

    assert (tmp_path / "custom.xlsx").exists()
    assert ctx.state["output_path"] == str(tmp_path / "custom.xlsx")


def test_build_output_sheets_rejects_empty_grouped_data(tmp_path):
    ctx = make_context(
        state={
            "grouped_data": {},
            "output_filename": "custom.xlsx",
        },
        config={
            "output_dir": str(tmp_path),
        },
    )

    with pytest.raises(BusinessException, match="No grouped data"):
        BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)


def test_build_output_sheets_rejects_output_filename_escape(tmp_path):
    ctx = make_context(
        state={
            "grouped_data": _grouped_data(),
            "output_filename": "../outside.xlsx",
        },
        config={
            "output_dir": str(tmp_path / "output"),
        },
    )

    with pytest.raises(SystemException, match="escapes output_dir"):
        BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)


def test_build_output_sheets_formats_subtotal_when_employee_name_matches_label(tmp_path):
    grouped_data = {
        "2024-01": [
            {"employee_name": "Subtotal", "date": "2024-01-05", "amount": 10.0, "country": "USA"},
            {"employee_name": "Zoe", "date": "2024-01-10", "amount": 15.5, "country": "UK"},
        ],
    }
    ctx = make_context(
        state={
            "grouped_data": grouped_data,
            "output_filename": "custom.xlsx",
        },
        config={
            "output_dir": str(tmp_path),
        },
    )

    BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)

    workbook = load_workbook(tmp_path / "custom.xlsx")
    assert workbook["2024-01"]["A2"].value == "Subtotal"
    assert workbook["2024-01"]["A2"].font.bold is False
    assert workbook["2024-01"]["A4"].value == "Subtotal"
    assert workbook["2024-01"]["A4"].font.bold is True


def test_build_output_sheets_cleans_temp_file_when_reload_fails(tmp_path, monkeypatch):
    ctx = make_context(
        state={
            "grouped_data": _grouped_data(),
            "output_filename": "custom.xlsx",
        },
        config={
            "output_dir": str(tmp_path),
        },
    )
    destination = tmp_path / "custom.xlsx"
    destination.write_text("previous output", encoding="utf-8")

    def fail_load_workbook(_path):
        raise OSError("cannot reload")

    monkeypatch.setattr(build_output_sheets, "load_workbook", fail_load_workbook)

    with pytest.raises(SystemException, match="Failed to build Excel workbook"):
        BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)

    assert destination.read_text(encoding="utf-8") == "previous output"
    assert list(tmp_path.glob(".custom.xlsx.*.tmp")) == []


def test_build_output_sheets_cleans_temp_file_when_replace_fails(tmp_path, monkeypatch):
    ctx = make_context(
        state={
            "grouped_data": _grouped_data(),
            "output_filename": "custom.xlsx",
        },
        config={
            "output_dir": str(tmp_path),
        },
    )
    destination = tmp_path / "custom.xlsx"
    destination.write_text("previous output", encoding="utf-8")

    def fail_replace(_src, _dst):
        raise OSError("cannot replace")

    monkeypatch.setattr(os, "replace", fail_replace)

    with pytest.raises(SystemException, match="Failed to build Excel workbook"):
        BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)

    assert destination.read_text(encoding="utf-8") == "previous output"
    assert list(tmp_path.glob(".custom.xlsx.*.tmp")) == []


def test_build_output_sheets_preserves_destination_when_fsync_fails(tmp_path, monkeypatch):
    ctx = make_context(
        state={
            "grouped_data": _grouped_data(),
            "output_filename": "custom.xlsx",
        },
        config={
            "output_dir": str(tmp_path),
        },
    )
    destination = tmp_path / "custom.xlsx"
    destination.write_text("previous output", encoding="utf-8")

    def fail_fsync(_descriptor):
        raise OSError("cannot fsync")

    monkeypatch.setattr(os, "fsync", fail_fsync)

    with pytest.raises(SystemException, match="Failed to build Excel workbook"):
        BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)

    assert destination.read_text(encoding="utf-8") == "previous output"
    assert list(tmp_path.glob(".custom.xlsx.*.tmp")) == []


def test_build_output_sheets_rejects_empty_employee_name(tmp_path):
    ctx = make_context(
        state={
            "grouped_data": {
                "2024-01": [
                    {"employee_name": "", "date": "2024-01-05", "amount": 10.0, "country": "USA"},
                ],
            },
            "output_filename": "custom.xlsx",
        },
        config={
            "output_dir": str(tmp_path),
        },
    )

    with pytest.raises(BusinessException, match="empty employee name"):
        BuildOutputSheets(name="build_output_sheets", execution_order=1).execute(ctx)
