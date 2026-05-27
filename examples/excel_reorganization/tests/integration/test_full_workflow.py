from __future__ import annotations

from openpyxl import load_workbook

from skills import BuildOutputSheets, GroupByMonth, LoadSalesData, VerifyOutput
from tests.conftest import make_context


def test_full_excel_reorganization_workflow(tmp_path):
    csv_path = tmp_path / "sales.csv"
    output_dir = tmp_path / "output"
    csv_path.write_text(
        "employee_name,date,amount,country\n"
        "Zoe,2024-01-10,15.50,UK\n"
        "Alice,2024-01-05,10.00,USA\n"
        "Bob,2024-02-01,7.00,Canada\n",
        encoding="utf-8",
    )
    ctx = make_context(
        {
            "csv_path": str(csv_path),
            "output_dir": str(output_dir),
            "output_filename": "sales_report_{month}.xlsx",
        }
    )

    LoadSalesData(name="load_sales_data", execution_order=1).execute(ctx)
    GroupByMonth(name="group_by_month", execution_order=2).execute(ctx)
    BuildOutputSheets(name="build_output_sheets", execution_order=3).execute(ctx)
    VerifyOutput(name="verify_output", execution_order=4).execute(ctx)

    output_path = output_dir / "sales_report_2024-01.xlsx"
    workbook = load_workbook(output_path)

    assert workbook.sheetnames == ["2024-01", "2024-02"]
    assert [row[0] for row in workbook["2024-01"].iter_rows(min_row=2, max_row=3, values_only=True)] == [
        "Alice",
        "Zoe",
    ]
    assert workbook["2024-01"]["C4"].value == 25.5
    assert workbook["2024-01"]["A1"].font.bold is True
