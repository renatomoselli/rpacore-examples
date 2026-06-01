"""
Unit tests for setup.py skills (OpenChallengePage, DownloadInputData, StartChallenge).

These tests use mocked browser objects to avoid requiring actual Playwright.
"""

import io
from pathlib import Path
from unittest.mock import Mock, patch

import openpyxl
import pytest
from rpacore import ProcessContext, Transaction, SystemException

from skills.setup import (
    OpenChallengePage,
    DownloadInputData,
    StartChallenge,
    DEFAULT_XLSX_URL,
)


class TestOpenChallengePage:
    """Test the OpenChallengePage skill."""

    def setup_method(self):
        """Set up test fixtures."""
        self.mock_pw = Mock()
        self.mock_browser = Mock()
        self.mock_page = Mock()
        self.mock_tx = Mock(spec=Transaction, reference="open-challenge-page")

    @patch("skills.setup.sync_playwright")
    def test_opens_challenge_page(self, mock_sync_pw):
        """Test that OpenChallengePage opens the challenge page."""
        mock_sync_pw_instance = Mock()
        mock_sync_pw.return_value = mock_sync_pw_instance
        mock_sync_pw_instance.start.return_value = self.mock_pw
        self.mock_pw.chromium.launch.return_value = self.mock_browser
        self.mock_browser.new_page.return_value = self.mock_page

        skill = OpenChallengePage("open_challenge_page", 1)
        ctx = ProcessContext(transaction=self.mock_tx, data={})
        skill.execute(ctx)

        # Verify launch was called with headless=True by default
        self.mock_pw.chromium.launch.assert_called_once_with(headless=True)
        self.mock_browser.new_page.assert_called_once()
        self.mock_page.goto.assert_called_with("https://www.rpachallenge.com/", timeout=30_000)
        assert ctx.data["page"] == self.mock_page
        assert ctx.data["_pw"] == self.mock_pw

    @patch("skills.setup.sync_playwright")
    def test_opens_challenge_page_headed(self, mock_sync_pw):
        """Test that OpenChallengePage opens with headed=False when config says so."""
        mock_sync_pw_instance = Mock()
        mock_sync_pw.return_value = mock_sync_pw_instance
        mock_sync_pw_instance.start.return_value = self.mock_pw
        self.mock_pw.chromium.launch.return_value = self.mock_browser
        self.mock_browser.new_page.return_value = self.mock_page

        skill = OpenChallengePage("open_challenge_page", 1)
        ctx = ProcessContext(transaction=self.mock_tx, data={})
        ctx.config = {"headless": "false"}
        skill.execute(ctx)

        # Verify launch was called with headless=False
        self.mock_pw.chromium.launch.assert_called_once_with(headless=False)

    @patch("skills.setup.sync_playwright")
    def test_raises_system_exception_on_launch_failure(self, mock_sync_pw):
        """Test that browser launch failure raises SystemException."""
        mock_sync_pw_instance = Mock()
        mock_sync_pw.return_value = mock_sync_pw_instance
        mock_sync_pw_instance.start.return_value = self.mock_pw
        self.mock_pw.chromium.launch.side_effect = Exception("Launch failed")

        skill = OpenChallengePage("open_challenge_page", 1)
        ctx = ProcessContext(transaction=self.mock_tx, data={})

        with pytest.raises(SystemException) as exc_info:
            skill.execute(ctx)
        assert "launch" in str(exc_info.value).lower()
        # pw.stop may be called multiple times in cleanup, just verify it was called
        self.mock_pw.stop.assert_called()

    @patch("skills.setup.sync_playwright")
    def test_raises_system_exception_on_navigation_failure(self, mock_sync_pw):
        """Test that navigation failure raises SystemException."""
        mock_sync_pw_instance = Mock()
        mock_sync_pw.return_value = mock_sync_pw_instance
        mock_sync_pw_instance.start.return_value = self.mock_pw
        self.mock_pw.chromium.launch.return_value = self.mock_browser
        self.mock_browser.new_page.return_value = self.mock_page
        self.mock_page.goto.side_effect = Exception("Navigation failed")

        skill = OpenChallengePage("open_challenge_page", 1)
        ctx = ProcessContext(transaction=self.mock_tx, data={})

        with pytest.raises(SystemException) as exc_info:
            skill.execute(ctx)
        assert "Failed to load page" in str(exc_info.value)


class TestDownloadInputData:
    """Test the DownloadInputData skill."""

    def setup_method(self):
        """Set up test fixtures."""
        self.mock_page = Mock()
        self.mock_tx = Mock(spec=Transaction, reference="download-input-data")
        self.mock_ctx = ProcessContext(
            transaction=self.mock_tx,
            data={"page": self.mock_page}
        )

    def _make_excel_bytes(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name", "Company Name",
                    "Role in Company", "Address", "Email", "Phone Number"])
        ws.append(["John", "Doe", "ACME", "Engineer",
                    "123 Main St", "john@example.com", "555-1234"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def test_downloads_and_parses_excel(self):
        """Test that DownloadInputData downloads and parses Excel."""
        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve:
            mock_urlretrieve.return_value = ("/tmp/test.xlsx", Mock())
            with patch("skills.setup.openpyxl.load_workbook") as mock_load_wb:
                mock_wb = Mock()
                mock_wb.active = Mock()
                mock_wb.active.iter_rows.return_value = iter([
                    ("First Name", "Last Name", "Company Name",
                     "Role in Company", "Address", "Email", "Phone Number"),
                    ("John", "Doe", "ACME", "Engineer",
                     "123 Main St", "john@example.com", "555-1234"),
                ])
                mock_load_wb.return_value = mock_wb

                skill = DownloadInputData("download_input_data", 1)
                skill.execute(self.mock_ctx)

                mock_urlretrieve.assert_called_once()
                mock_load_wb.assert_called_once()

    def test_uses_configured_xlsx_url(self):
        """Test that DownloadInputData uses custom xlsx_url from config."""
        self.mock_ctx.config = {"xlsx_url": "http://custom-url.com/data.xlsx"}

        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve:
            mock_urlretrieve.return_value = ("/tmp/test.xlsx", Mock())
            with patch("skills.setup.openpyxl.load_workbook") as mock_load_wb:
                mock_wb = Mock()
                mock_wb.active = Mock()
                mock_wb.active.iter_rows.return_value = iter([
                    ("First Name", "Last Name", "Company Name",
                     "Role in Company", "Address", "Email", "Phone Number"),
                    ("John", "Doe", "ACME", "Engineer",
                     "123 Main St", "john@example.com", "555-1234"),
                ])
                mock_load_wb.return_value = mock_wb

                skill = DownloadInputData("download_input_data", 1)
                skill.execute(self.mock_ctx)

                mock_urlretrieve.assert_called_once()
                args, _ = mock_urlretrieve.call_args
                assert args[0] == "http://custom-url.com/data.xlsx"

    def test_falls_back_to_browser_download_on_direct_failure(self):
        """Test that DownloadInputData falls back to browser download."""
        from unittest.mock import MagicMock
        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve:
            mock_urlretrieve.side_effect = Exception("Download failed")
            with patch("skills.setup.openpyxl.load_workbook") as mock_load_wb, \
                 patch("pathlib.Path.unlink"), \
                 patch("tempfile.mkstemp", return_value=(123, "/tmp/test.xlsx")):
                mock_wb = Mock()
                mock_wb.active = Mock()
                mock_wb.active.iter_rows.return_value = iter([
                    ("First Name", "Last Name", "Company Name",
                     "Role in Company", "Address", "Email", "Phone Number"),
                    ("John", "Doe", "ACME", "Engineer",
                     "123 Main St", "john@example.com", "555-1234"),
                ])
                mock_load_wb.return_value = mock_wb

                mock_download = Mock()
                mock_download.save_as = Mock()
                state = {"expecting_download": False}

                def enter_download_context():
                    state["expecting_download"] = True
                    return mock_context_mgr

                def exit_download_context(*_args):
                    state["expecting_download"] = False
                    return None

                def click_download(*_args, **_kwargs):
                    assert state["expecting_download"] is True

                # Build a proper context manager for expect_download
                mock_context_mgr = MagicMock()
                mock_context_mgr.__enter__ = Mock(side_effect=enter_download_context)
                mock_context_mgr.__exit__ = Mock(side_effect=exit_download_context)
                mock_context_mgr.value = mock_download
                # Set expect_download as a method on the page mock
                self.mock_page.expect_download = Mock(return_value=mock_context_mgr)
                self.mock_page.click.side_effect = click_download

                self.mock_ctx.config = {"xlsx_url": "http://example.com/file.xlsx"}

                skill = DownloadInputData("download_input_data", 1)
                skill.execute(self.mock_ctx)

                # Should have tried urlretrieve first, then browser download
                assert mock_urlretrieve.call_count >= 1
                # Should have clicked download button
                self.mock_page.click.assert_called()
                # Should have called expect_download
                self.mock_page.expect_download.assert_called_once()

    def test_raises_system_exception_on_parse_failure(self):
        """Test that parse failure raises SystemException."""
        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve:
            mock_urlretrieve.return_value = ("/tmp/test.xlsx", Mock())
            with patch("skills.setup.openpyxl.load_workbook") as mock_load_wb:
                mock_wb = Mock()
                mock_wb.active = Mock()
                mock_wb.active.iter_rows.side_effect = Exception("Invalid Excel")
                mock_load_wb.return_value = mock_wb

                skill = DownloadInputData("download_input_data", 1)

                with pytest.raises(SystemException) as exc_info:
                    skill.execute(self.mock_ctx)

                assert "Failed to parse Excel file" in str(exc_info.value)

    def test_raises_system_exception_on_empty_data(self):
        """Test that empty Excel data fails setup before the browser is driven."""
        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve:
            mock_urlretrieve.return_value = ("/tmp/test.xlsx", Mock())
            with patch("skills.setup.openpyxl.load_workbook") as mock_load_wb:
                mock_wb = Mock()
                mock_wb.active = Mock()
                # Return headers but no data rows — triggers the "no data rows" path
                mock_wb.active.iter_rows.return_value = iter([
                    ("First Name", "Last Name", "Company Name",
                     "Role in Company", "Address", "Email", "Phone Number"),
                ])
                mock_load_wb.return_value = mock_wb

                skill = DownloadInputData("download_input_data", 1)

                with pytest.raises(SystemException) as exc_info:
                    skill.execute(self.mock_ctx)

                assert "no data rows" in str(exc_info.value).lower()

    def test_raises_system_exception_on_missing_row_values(self):
        """Test invalid row data fails during setup before browser submission."""
        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve:
            mock_urlretrieve.return_value = ("/tmp/test.xlsx", Mock())
            with patch("skills.setup.openpyxl.load_workbook") as mock_load_wb:
                mock_wb = Mock()
                mock_wb.active = Mock()
                mock_wb.active.iter_rows.return_value = iter([
                    ("First Name", "Last Name", "Company Name",
                     "Role in Company", "Address", "Email", "Phone Number"),
                    ("John", "", "ACME", "Engineer",
                     "123 Main St", "john@example.com", "555-1234"),
                ])
                mock_load_wb.return_value = mock_wb

                skill = DownloadInputData("download_input_data", 1)

                with pytest.raises(SystemException) as exc_info:
                    skill.execute(self.mock_ctx)

                assert "missing required value" in str(exc_info.value)


class TestStartChallenge:
    """Test the StartChallenge skill."""

    def setup_method(self):
        """Set up test fixtures."""
        self.mock_page = Mock()
        self.mock_tx = Mock(spec=Transaction, reference="start-challenge")
        self.mock_ctx = ProcessContext(
            transaction=self.mock_tx,
            data={"page": self.mock_page}
        )

    def test_clicks_start_button(self):
        """Test that StartChallenge clicks the start button."""
        # query_selector returns None → button exists, needs clicking
        self.mock_page.query_selector.return_value = Mock()

        skill = StartChallenge("start_challenge", 1)
        skill.execute(self.mock_ctx)

        # Verify click uses the has-text selector
        self.mock_page.click.assert_called_with('button:has-text("START")', timeout=10_000)

    def test_waits_for_form_appearance(self):
        """Test that StartChallenge waits for form fields to appear."""
        self.mock_page.query_selector.return_value = Mock()

        skill = StartChallenge("start_challenge", 1)
        skill.execute(self.mock_ctx)

        # Verify wait_for_function was called with the rpa1-field check
        self.mock_page.wait_for_function.assert_called_once()

    def test_skips_when_form_already_visible(self):
        """Test that StartChallenge returns early if form is already visible."""
        # query_selector returns None → START button not on page
        self.mock_page.query_selector.return_value = None

        skill = StartChallenge("start_challenge", 1)
        skill.execute(self.mock_ctx)

        # Should not click or wait for anything
        self.mock_page.click.assert_not_called()
        self.mock_page.wait_for_function.assert_not_called()

    def test_raises_system_exception_on_start_failure(self):
        """Test that start failure raises SystemException."""
        self.mock_page.query_selector.side_effect = Exception("Start failed")

        skill = StartChallenge("start_challenge", 1)

        with pytest.raises(SystemException) as exc_info:
            skill.execute(self.mock_ctx)

        assert "Failed to start the challenge" in str(exc_info.value)
