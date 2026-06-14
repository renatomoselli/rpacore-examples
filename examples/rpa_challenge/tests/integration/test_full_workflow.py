from __future__ import annotations

"""
Integration tests for the full RPA workflow.

These tests verify that all skills work together correctly.
"""

import io
import sys
import urllib.error
from pathlib import Path
from unittest.mock import Mock, patch

import openpyxl
import pytest

# Add the parent directory to the path
sys.path.insert(0, str(Path(__file__).parent.parent))

from rpacore import ProcessContext, Transaction, SystemException
from skills.setup import (
    OpenChallengePage,
    DownloadInputData,
    StartChallenge,
    DEFAULT_XLSX_URL,
)
from skills.row import FillRow, SubmitRow
from skills.score import RecordScore

pytestmark = pytest.mark.integration


class TestFullWorkflow:
    """Test the full workflow from start to finish."""

    def _create_test_excel_bytes(self):
        """Create test Excel data as bytes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name", "Company Name",
                    "Role in Company", "Address", "Email", "Phone Number"])
        ws.append(["John", "Doe", "ACME Corp", "Software Engineer",
                    "123 Main St", "john@example.com", "555-0100"])
        ws.append(["Jane", "Smith", "Tech Solutions", "Product Manager",
                    "456 Oak Ave", "jane@example.com", "555-0200"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def test_download_input_data_populates_state(self):
        """Test that downloaded workbook rows are stored in durable state."""
        mock_tx = Mock(spec=Transaction, reference="test-workflow", state={})

        # Mock the browser and download
        mock_page = Mock()
        mock_pw = Mock()
        mock_browser = Mock()

        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve, \
             patch("skills.setup.openpyxl.load_workbook") as mock_load_wb:

            mock_urlretrieve.return_value = ("test.xlsx", Mock())

            # Mock Excel parsing
            mock_wb = Mock()
            mock_wb.active = Mock()
            mock_wb.active.iter_rows.return_value = iter([
                ("First Name", "Last Name", "Company Name",
                 "Role in Company", "Address", "Email", "Phone Number"),
                ("John", "Doe", "ACME Corp", "Software Engineer",
                 "123 Main St", "john@example.com", "555-0100"),
                ("Jane", "Smith", "Tech Solutions", "Product Manager",
                 "456 Oak Ave", "jane@example.com", "555-0200"),
            ])
            mock_load_wb.return_value = mock_wb

            ctx = ProcessContext(
                transaction=mock_tx,
                resources={"page": mock_page, "_pw": mock_pw},
                config={
                    "xlsx_url": "http://test.example.com/data.xlsx",
                    "xlsx_allowed_hosts": ["test.example.com"],
                },
            )

            skill2 = DownloadInputData("download_input_data", 2)
            skill2.execute(ctx)

            assert ctx.resources.get("page") == mock_page
            # Verify DownloadInputData populated durable state
            assert ctx.state.get("rows") is not None
            assert len(ctx.state["rows"]) == 2

    def test_workflow_uses_custom_config(self):
        """Test that workflow respects custom configuration."""
        mock_tx = Mock(spec=Transaction, reference="test-config", state={})
        mock_page = Mock()
        mock_pw = Mock()

        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve, \
             patch("pathlib.Path.unlink"), \
             patch("tempfile.mkstemp", return_value=(123, "test.xlsx")), \
             patch("skills.setup.openpyxl.load_workbook") as mock_load_wb:
            mock_urlretrieve.side_effect = urllib.error.URLError("Network error")
            # Mock the browser download fallback via expect_download context manager
            mock_download = Mock()
            mock_download.save_as = Mock()
            mock_context_mgr = Mock()
            mock_context_mgr.value = mock_download  # dl_info.value is the download object
            mock_context_mgr.__enter__ = Mock(return_value=mock_context_mgr)
            mock_context_mgr.__exit__ = Mock(return_value=None)
            mock_page.expect_download = Mock(return_value=mock_context_mgr)
            mock_page.wait_for_selector = Mock()
            mock_page.click = Mock()
            mock_wb = Mock()
            mock_wb.active = Mock()
            mock_wb.active.iter_rows.return_value = iter([
                ("First Name", "Last Name", "Company Name",
                 "Role in Company", "Address", "Email", "Phone Number"),
                ("John", "Doe", "ACME", "Engineer",
                 "123 Main St", "john@example.com", "555-1234"),
            ])
            mock_load_wb.return_value = mock_wb

            ctx = ProcessContext(
                transaction=mock_tx,
                resources={"page": mock_page, "_pw": mock_pw},
                config={
                    "xlsx_url": "http://custom-test-url.com/data.xlsx",
                    "xlsx_allowed_hosts": ["custom-test-url.com"],
                },
            )

            # Skill should use custom URL from config
            skill = DownloadInputData("download_input_data", 2)
            skill.execute(ctx)

            # Verify urlretrieve was called (it fails, then fallback happens)
            assert mock_urlretrieve.call_count >= 1
            # Verify browser download fallback was used
            mock_page.expect_download.assert_called_once()
            mock_page.click.assert_called_once()
            mock_download.save_as.assert_called_once()

    def test_workflow_handles_parse_errors(self):
        """Test that workflow handles Excel parse errors."""
        mock_tx = Mock(spec=Transaction, reference="test-errors", state={})
        mock_page = Mock()

        with patch("skills.setup.urllib.request.urlretrieve") as mock_urlretrieve, \
             patch("skills.setup.openpyxl.load_workbook") as mock_load_wb:

            mock_urlretrieve.return_value = ("test.xlsx", Mock())
            mock_wb = Mock()
            mock_wb.active = Mock()
            mock_wb.active.iter_rows.side_effect = Exception("Parse error")
            mock_load_wb.return_value = mock_wb

            ctx = ProcessContext(
                transaction=mock_tx,
                resources={"page": mock_page},
                config={
                    "xlsx_url": "http://test.example.com/data.xlsx",
                    "xlsx_allowed_hosts": ["test.example.com"],
                },
            )

            # Download should raise an exception on parse failure
            with pytest.raises(SystemException) as exc_info:
                skill = DownloadInputData("download_input_data", 2)
                skill.execute(ctx)

            assert "Failed to parse Excel file" in str(exc_info.value)
